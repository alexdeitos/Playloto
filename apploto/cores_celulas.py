from turtle import color
from openpyxl.styles import *

def yellow1():
    yellow = PatternFill(
        start_color='ffff00',
        end_color='000000',
        fill_type='lightVertical'
    )
    return yellow

def black1():
    black = PatternFill(
        start_color='000000',
        end_color='000000',
        fill_type='solid'
    )
    return black

def red1():
    red = PatternFill(
        start_color='cd5c5c',
        end_color='cd5c5c',
        fill_type='solid'
    )
    return red

def ligth_gray1():
    ligth_gray = PatternFill(
        start_color='808080',
        end_color='808080',
        fill_type='solid'
    )
    return ligth_gray

def thin_border1():
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    return thin_border

def fonte1():
    fonte = Font(
        bold=True,
        size=14,
        italic=False,
        color="000000"
    )
    return fonte

def alinhamento1():
    alinhamento = Alignment(
        horizontal='center'
    )
    return alinhamento


def blue():
    yellow = PatternFill(
        start_color='0000ff',
        end_color='0000ff',
        fill_type='lightVertical'
    )
    return yellow


def white():
    white = PatternFill(
        start_color='ffffff',
        end_color='ffffff',
        fill_type='lightVertical'
    )
    return white