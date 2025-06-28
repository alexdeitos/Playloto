#from itertools import count
from operator import length_hint
from django.shortcuts import render,redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.template import Library;register = Library()
#from .models import Sorteio, MeusJogos
#from django.core.paginator import Paginator
from datetime import date, timedelta
from random import randint
#from openpyxl import Workbook
#from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from .cores_celulas import *
from .forms import FixasForm, Valida
#from django.views.generic import ListView
from .forms import Valida
import csv
#import io
#from django.http import FileResponse
#from django.template import Context
from django.template.loader import get_template
import openpyxl
from datetime import datetime
from .models import Sorteio


# Create your views here.
def index(request):
    last = Sorteio.objects.last()  # Obtém o último sorteio do banco de dados

    # Se não houver sorteios, redirecione para a página de importação
    if not last:
        return redirect('apploto:importar_sorteios')

    last = Sorteio.objects.all().latest()
    return render(request, 'index.html', {'last' : last})

@login_required
def tabelaMovimento(request):
    today_date = date.today()
    td = timedelta(15)
    data =  today_date - td
    last = Sorteio.objects.all().latest()
    last_concurso = last.concurso
    last_data = last.data_sorteio
    query_set = Sorteio.objects.values_list('B1','B2','B3','B4','B5','B6','B7','B8','B9','B10','B11','B12','B13','B14','B15').filter(concurso__gte=(last_concurso-15)).order_by('concurso')
    #query_set = Sorteio.objects.values_list('B1','B2','B3','B4','B5','B6','B7','B8','B9','B10','B11','B12','B13','B14','B15').order_by('concurso')
    query_set1 = Sorteio.objects.values_list('B1','B2','B3','B4','B5','B6','B7','B8','B9','B10','B11','B12','B13','B14','B15').filter(data_sorteio__gte=data).order_by('-concurso')
    result = []
    result1 = []
    # lista_games agora é uma lista de jogos de 15 dezenas cada jogo
    # contador = 0 
    
    for c in query_set:
        result.append(c)
    for c in query_set1:
        result1.append(c)
    vetJogo = ['x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x',
               'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x']

    vetJogos = list()
    vetJogos1 = list()
    for sorteio in result:
        for dezena in sorteio:
            vetJogo[dezena-1] = dezena
        vetJogos.append(vetJogo[:])
        vetJogo.clear()
        vetJogo = ['x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x',
                   'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x']
    for sorteio in result1:
        for dezena in sorteio:
            vetJogo[dezena-1] = dezena
        vetJogos1.append(vetJogo[:])
        vetJogo.clear()
        vetJogo = ['x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x',
                   'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x']
    ciclo = []
    countCiclo = 0
    for jogo in vetJogos:
        for dezena in jogo:
            if dezena in (range(1,25)):
                if dezena not in ciclo:
                    ciclo.append(dezena)
                if len(ciclo) == 25:
                    countCiclo += 1
                    context = {'countCiclo' : countCiclo, }

    context = {
        'qtd_jogos' : len(vetJogos),
        'vetJogos': vetJogos,
        'vetJogos1' : vetJogos1,
        'last_concurso' : last_concurso, 
        'nome_fibo' : 'Fibonacci',
        'data_sorteio':last_data,
    }

    return render(request, 'TabelaMovimento.html',context)

@login_required
def sorteia(request):
    
    while True:
        #variáveis de validação
        countRepetidas = 0
        impar = fibo = primo = multiplo = moldura = soma = countfixas = 0
        vetorPrimos = [2, 3, 5, 7, 11, 13, 17, 19, 23]
        vetorMoldura = [1, 2, 3, 4, 5, 6, 10, 11, 15, 16, 20, 21, 22, 23, 24, 25]
        vetorMultiplos = [3, 6, 9, 12, 15, 18, 21, 24]
        vetorFibonacci = [2, 3, 5, 8, 13, 21]
        last_game = list()
        novo_sorteio = list()
        form = FixasForm()
        fixas = ''
        lista_fixas =[]
        #adiciona o último sorteio ao last
        last = Sorteio.objects.all().latest()
        
        # adiciona os valores de last a lista last_games
        last_game.append(last.B1)
        last_game.append(last.B2)
        last_game.append(last.B3)
        last_game.append(last.B4)
        last_game.append(last.B5)
        last_game.append(last.B6)
        last_game.append(last.B7)
        last_game.append(last.B8)
        last_game.append(last.B9)
        last_game.append(last.B10)
        last_game.append(last.B11)
        last_game.append(last.B12)
        last_game.append(last.B13)
        last_game.append(last.B14)
        last_game.append(last.B15)
        
        while True:
            aleatorio = randint(1,25)
            if aleatorio not in novo_sorteio:
                novo_sorteio.append(aleatorio)
                if len(novo_sorteio) == 15:
                    break
        
        if request.method == 'POST':
            # Create a form instance and populate it with data from the request (binding):
            form = FixasForm(request.POST)
            # Check if the form is valid:
            if form.is_valid():
                # process the data in form.cleaned_data as required (here we just write it to the model due_back field)
                fixas = form.cleaned_data['fixas']
                fixas = fixas.split(',')
                for f in fixas:
                    lista_fixas.append(int(f))
        
        # realiza um count das variáveis para retornar para o usuário 
        for n in novo_sorteio:
            soma += n
            
            if n in last_game:
                countRepetidas += 1
            if n % 2 == 1:
                impar += 1
            if n in vetorPrimos:
                primo += 1
            if n in vetorFibonacci:
                fibo += 1
            if n in vetorMoldura:
                moldura += 1
            if n in vetorMultiplos:
                multiplo += 1
            if n in lista_fixas:
                countfixas += 1
        '''soma > 190 and soma < 200 and'''       
        if  soma > 185 and soma < 208 and countfixas == len(lista_fixas) and impar == 7 and fibo == 4 and primo == 5:
            #break
        # adiciona o resultados das variáveis a uma variável que será enviada ao cliente
        
            context = {
                "novo_sorteio" : sorted(novo_sorteio),
                'impar': impar,
                'fibo': fibo,
                'primo': primo,
                'moldura': moldura,
                'multiplo': multiplo,
                'countRepetidas': countRepetidas,
                'soma' : soma,
                'form' : form,
                'fixas': fixas,
                'lista_fixas' : lista_fixas,
            }
            break
    
    return render(request, 'sorteia.html', context)

'''
@login_required
def planilha(request):
    
    dezenas = [i for i in range(1, 26)] 
    
    query_set1 = Sorteio.objects.values_list('B1','B2','B3','B4','B5','B6','B7','B8','B9','B10','B11','B12','B13','B14','B15')
    result1 = []
        
    for d in query_set1:
        result1.append(d)

    # csv file 
    f = open('games.csv', 'w')
    f = HttpResponse(
        content_type='text/csv',
        headers={'content-disposition': 'attachment; filename=games.csv'},
    )
    writer = csv.writer(f)
    concurso = 1
    count = 0 
    header = ['concurso', 'd1', 'd2', 'd3', 'd4', 'd5', 'd6',
        'd7', 'd8', 'd9', 'd10', 'd11', 'd12', 'd13', 'd14',
        'd15', 'faltam', 'ciclo']
    dezenas = [i for i in range(1, 26)]
    text = ["sep=,"]
    writer.writerow(text)
    writer.writerow(header)
    
    dados = []
    dadosfinais = []
    for j in result1:
        for n in j:
            dados.append(n)
        dadosfinais.append(dados[:])
        dados.clear()
    
    for j in dadosfinais:
        if len(dezenas) == 0:
            dezenas = [i for i in range(1, 26)]
        for i in j:
            if i in dezenas:
                dezenas.remove(i)
                if len(dezenas) == 0:
                    count+= 1
        l1 = j[0]    
        l2 = j[1]
        l3 = j[2]
        l4 = j[3]
        l5 = j[4]
        l6 = j[5]
        l7 = j[6]
        l8 = j[7]
        l9 = j[8]
        l10 = j[9]
        l11 = j[10]
        l12 = j[11]
        l13 = j[12]
        l14 = j[13]
        l15 = j[14]
        if len(dezenas)==0:
            l = [concurso, l1, l2, l3, l4, l5, l6, l7, l8, l9,
            l10, l11, l12, l13, l14, l15, 'ciclo encerrado', count]
        else:
            l = [concurso, l1, l2, l3, l4, l5, l6, l7, l8, l9,
            l10, l11, l12, l13, l14, l15, dezenas]
        concurso += 1
        writer.writerow(l)
        
    f.close()    
    return f
'''

# correcao do chatgpt 2025-06-28
@login_required
def planilha(request):
    
    dezenas = list(range(1, 26))
    
    query_set1 = Sorteio.objects.values_list(
        'B1','B2','B3','B4','B5','B6','B7','B8','B9','B10',
        'B11','B12','B13','B14','B15'
    )
    
    response = HttpResponse(
        content_type='text/csv',
        headers={'content-disposition': 'attachment; filename=games.csv'},
    )
    writer = csv.writer(response)
    
    writer.writerow(["sep=,"])
    header = ['concurso'] + [f'd{i}' for i in range(1, 16)] + ['faltam', 'ciclo']
    writer.writerow(header)
    
    count = 0
    concurso = 1
    
    for j in query_set1:
        if len(dezenas) == 0:
            dezenas = list(range(1, 26))
        
        # Remove dezenas presentes no concurso
        for i in j:
            if i in dezenas:
                dezenas.remove(i)
        
        # Verifica ciclo encerrado
        if len(dezenas) == 0:
            count += 1
            faltam_val = 'ciclo encerrado'
            ciclo_val = count
            dezenas = []  # para garantir
        else:
            faltam_val = str(dezenas)
            ciclo_val = ''
        
        linha = [concurso] + list(j) + [faltam_val, ciclo_val]
        writer.writerow(linha)
        concurso += 1
    
    return response

@login_required
# defina a função para obter os dados da tabela
def get_dados_tabela():
    # substitua esta lógica de exemplo pela lógica real para obter seus dados
    # por exemplo, você pode consultar o modelo sorteio para obter os dados
    dados = Sorteio.objects.all()  # consulta para obter todos os objetos sorteio
    return dados

@login_required
def exportar_tabela_excel(request):
    # recupere todos os jogos do banco de dados
    jogos = Sorteio.objects.all().order_by('-concurso')

    # crie um contexto para passar os jogos para o template do excel
    context = {
        'dados': jogos,
    }

    # renderize o template html que contém a tabela em excel
    template = get_template('tabela_excel.html')
    html = template.render(context)

    # crie um arquivo temporário para armazenar o pdf
    result = HttpResponse(content_type='application/vnd.ms-excel')
    result['content-disposition'] = 'attachment; filename="tabela_excel.xls"'

    # converta o html para pdf usando a biblioteca xhtml2pdf
    #pdf = pisa.pisadocument(html, result)

    # verifique se a conversão para pdf foi bem-sucedida
    #if not pdf.err:
    #    return result

    return HttpResponse('erro ao gerar o arquivo excel.', content_type='text/plain')

@login_required
def descubra(request):
    # recuperar o último sorteio
    last = Sorteio.objects.all().latest()

    # inicializar variáveis
    form = Valida()
    ultimo_sorteio = [
        last.B1, last.B2, last.B3, last.B4, last.B5,
        last.B6, last.B7, last.B8, last.B9, last.B10,
        last.B11, last.B12, last.B13, last.B14, last.B15
    ]
    acertos = 0

    # processar o formulário se for um post
    if request.method == 'post':
        form = Valida(request.post)
        if form.is_valid():
            jogo_digitado = form.cleaned_data['jogo']
            jogo_digitado = [int(numero) for numero in jogo_digitado.split(',')]

            # comparar os números do jogo digitado com o último sorteio
            acertos = sum(1 for numero in jogo_digitado if numero in ultimo_sorteio)

    context = {
        "last": last,
        "acertos": acertos,
        "form": form,
    }

    return render(request, 'descubra.html', context)

def scrapping(request):
    return render(request, 'resultados.html', context={})

def importar_sorteios(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        if file.name.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file)
            worksheet = workbook.active

            # Iterar pelas linhas a partir da segunda (ignorando o cabeçalho)
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                # Checar se a linha é vazia
                if row[0] is not None:
                    sorteio_id = row[0]
                    data_sorteio = row[1]  # Data já está no formato datetime
                    numeros = row[2:17]  # Coluna B1 até B15
                    qtd_ganhadores_15 = row[17]

                    # Criar um novo objeto Sorteio e salvá-lo no banco de dados
                    sorteio = Sorteio(
                        concurso=sorteio_id, 
                        data_sorteio=data_sorteio,
                        B1=numeros[0], B2=numeros[1], B3=numeros[2], B4=numeros[3],
                        B5=numeros[4], B6=numeros[5], B7=numeros[6], B8=numeros[7],
                        B9=numeros[8], B10=numeros[9], B11=numeros[10], B12=numeros[11],
                        B13=numeros[12], B14=numeros[13], B15=numeros[14],
                        qtd_ganhadores_15=qtd_ganhadores_15
                    )
                    sorteio.save()

            # Obter o último sorteio inserido para exibição
            last = Sorteio.objects.all().latest()
            return render(request, 'index.html', {'last': last})
        else:
            return HttpResponse("Por favor, selecione um arquivo .xlsx.")
    return render(request, 'importar_sorteios.html')

def todos_sorteios_mega_sena(request):
    mega_sena_sorteios = Sorteio.objects.all()
    return render(request, 'sorteios_mega_sena.html', {'sorteios': mega_sena_sorteios})

